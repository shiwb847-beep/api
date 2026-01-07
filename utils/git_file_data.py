import os

from openpyxl import load_workbook
from config import Config

#获取xlsx数据
def get_xlsx_data(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    max_column = sheet.max_column
    data = []
    for i in range(2, max_row + 1):
        row_data = {}
        for j in range(1, max_column + 1):
            row_data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        data.append(row_data)
    return  data